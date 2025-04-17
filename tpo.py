from flask import Blueprint, render_template, request, session, redirect, url_for
from openpyxl import load_workbook

tpo_bp = Blueprint('tpo', __name__)

@tpo_bp.route('/dashboard')
def dashboard():
    if 'user_id' not in session or session.get('role') != 'tpo':
        return redirect(url_for('auth.login'))
    return render_template('tpo/dashboard.html')

@tpo_bp.route('/filter-students', methods=['GET', 'POST'])
def filter_students():
    if 'user_id' not in session or session.get('role') != 'tpo':
        return redirect(url_for('auth.login'))
    
    if request.method == 'POST':
        min_cgpa = float(request.form.get('min_cgpa', 0))
        branch = request.form.get('branch', '')
        skills = request.form.get('skills', '')
        
        wb = load_workbook('job_data.xlsx')
        ws = wb.active
        
        filtered_students = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            student_data = dict(zip(headers, row))
            
            if (float(student_data.get('CGPA', 0)) >= min_cgpa and \
               (not branch or student_data.get('Branch', '').lower() == branch.lower()) and \
               (not skills or all(skill.lower() in student_data.get('Skills', '').lower() 
                               for skill in skills.split(','))):
                filtered_students.append(student_data)
        
        return render_template('tpo/results.html', students=filtered_students)
    
    return render_template('tpo/filter.html')