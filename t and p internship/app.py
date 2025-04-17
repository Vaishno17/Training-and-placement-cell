from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Excel file paths
JOB_EXCEL_FILE = "job_data.xlsx"
MS_EXCEL_FILE = "ms_data.xlsx"
MBA_EXCEL_FILE = "mba_data.xlsx"

# Function to initialize Excel file with headers
def init_excel(file_path, headers):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_path)

# Initialize files with headers
init_excel(JOB_EXCEL_FILE, ["Name", "Email", "Phone", "Branch", "CGPA", "Skills", "Preferred Company"])
init_excel(MS_EXCEL_FILE, ["Name", "Email", "Phone", "Branch", "CGPA", "Skills", "Location", "Preferred University"])
init_excel(MBA_EXCEL_FILE, ["Name", "Email", "Phone", "Branch", "CGPA", "Skills", "Preferred B-School"])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/form/<path>')
def form_path(path):
    if path == 'job':
        return render_template('job_form.html')
    elif path == 'ms':
        return render_template('ms_form.html')
    elif path == 'mba':
        return render_template('mba_form.html')
    else:
        return "<h2>404 - Page Not Found</h2>"

@app.route('/submit_job', methods=['POST'])
def submit_job():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    branch = request.form['branch']
    cgpa = request.form['cgpa']
    skills = request.form['skills']
    company = request.form['company']

    wb = load_workbook(JOB_EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, phone, branch, cgpa, skills, company])
    wb.save(JOB_EXCEL_FILE)

    return "Job form submitted successfully!"

@app.route('/submit_ms', methods=['POST'])
def submit_ms():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    branch = request.form['branch']
    cgpa = request.form['cgpa']
    skills = request.form['skills']
    location = request.form['location']
    university = request.form['university']

    wb = load_workbook(MS_EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, phone, branch, cgpa, skills, location, university])
    wb.save(MS_EXCEL_FILE)

    return "MS form submitted successfully!"

@app.route('/submit_mba', methods=['POST'])
def submit_mba():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    branch = request.form['branch']
    cgpa = request.form['cgpa']
    skills = request.form['skills']
    bschool = request.form['bschool']

    wb = load_workbook(MBA_EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, phone, branch, cgpa, skills, bschool])
    wb.save(MBA_EXCEL_FILE)

    return "MBA form submitted successfully!"

if __name__ == '__main__':
    app.run(debug=True)
