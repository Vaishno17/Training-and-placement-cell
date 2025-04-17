from openpyxl import load_workbook, Workbook
from datetime import datetime

JOB_EXCEL_FILE = "job_data.xlsx"
ELIGIBILITY_FILE = "job_eligibility.xlsx"

# Eligibility criteria (can be modified)
MIN_CGPA = 7.0
REQUIRED_SKILLS = ["Python", "Java"]  # Example required skills
MAX_BACKLOGS = 2

def check_eligibility(student_data):
    """Check if student meets all eligibility criteria"""
    cgpa = float(student_data.get("CGPA", 0))
    skills = student_data.get("Skills", "").lower()
    backlogs = int(student_data.get("Backlogs", 0))
    
    # Check CGPA
    if cgpa < MIN_CGPA:
        return False, "CGPA too low"
    
    # Check skills
    student_skills = [skill.strip().lower() for skill in skills.split(",")]
    required_skills_present = all(
        any(req_skill.lower() in student_skill for student_skill in student_skills)
        for req_skill in REQUIRED_SKILLS
    )
    if not required_skills_present:
        return False, "Missing required skills"
    
    # Check backlogs
    if backlogs > MAX_BACKLOGS:
        return False, "Too many backlogs"
    
    return True, "Eligible"

def generate_eligibility_report():
    """Generate eligibility report for all students"""
    # Load job data
    wb = load_workbook(JOB_EXCEL_FILE)
    ws = wb.active
    
    # Create new workbook for eligibility
    new_wb = Workbook()
    new_ws = new_wb.active
    
    # Copy headers and add new columns
    headers = [cell.value for cell in ws[1]]
    headers.extend(["Eligibility", "Reason", "Date Checked"])
    new_ws.append(headers)
    
    # Iterate through data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        student_data = dict(zip(headers[:7], row))  # First 7 columns are original data
        
        # Check eligibility
        is_eligible, reason = check_eligibility(student_data)
        eligibility_status = "Eligible" if is_eligible else "Not Eligible"
        
        # Add to new worksheet
        new_row = list(row) + [eligibility_status, reason, datetime.now().strftime("%Y-%m-%d")]
        new_ws.append(new_row)
    
    # Save to new file
    new_wb.save(ELIGIBILITY_FILE)
    print(f"Eligibility report generated at {ELIGIBILITY_FILE}")

if __name__ == "__main__":
    generate_eligibility_report()